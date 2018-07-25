# -*- coding: utf-8 -*-
"""
Created on Tue Jul 17 10:01:41 2018

@author: test
"""
def cell(argument):
        switcher = {
            3: 'A30',
            4: "K30",
            5: "U30",
            6: 'AE30',
            7: 'AN30',
            8: 'AX30',
            9: 'A45',
            10: "K45",
            11: "U45",
            12: 'AE45',
            13: 'AN45',
            14: 'AX45',
            15: 'A60',
            16: "K60",
            17: "U60",
            18: 'AE60',
            19: 'AN60',
            20: 'AX60',
            21: 'A75',
            22: "K75",
            23: "U75",
            24: 'AE75',
            25: 'AN75',
            26: 'AX75',
            27: 'A90',
            28: "K90",
            29: "U90",
            30: 'AE90',
            31: 'AN90',
            32: 'AX90',
            33: 'A105',
            34: "K105",
            35: "U105",
            36: 'AE105',
            37: 'AN105',
            38: 'AX105',

        }
        return switcher.get(argument)



def scatchart(file_chart,vdd):
    import time
    import numpy as np
    from openpyxl.chart.layout import Layout, ManualLayout
    from openpyxl.chart import (
        ScatterChart,
        Reference,
        Series,
    )
    from openpyxl import load_workbook
    from datetime import date

    from openpyxl import Workbook
    from openpyxl.chart import (
            LineChart,
            Reference,
    )
    from openpyxl.chart.axis import DateAxis
    from excel_modf import switch_demo
    wb = load_workbook(filename=file_chart, read_only=False)
    ws = wb['summary_table']
    
    count = 0

#    for j in range(0,len(vdd)):
#        
#        chart = ScatterChart()
#        chart.style = 1
#        chart.title = 'Settling_times_VDD'+'_'+ str(vdd[j])+' V'
#        chart.scatterStyle = 'marker'
#        chart.x_axis.title = 'TEMP'
#        chart.y_axis.title = 'Settling times'
##        chart.layout= Layout(
##                manualLayout=ManualLayout(
##                x=0, y=0,
##                h=0.7, w=0.7,
##                )
##        )
#        
#        xvalues = Reference(ws, min_col=1, min_row=4, max_row=15)
#        #for j in range(0,54,18):
#        for i in range(2+count, 20+count):
#            values = Reference(ws, min_col=i, min_row=1, max_row=13)
#            series = Series(values, xvalues, title_from_data=True)
#            chart.series.append(series)
#        #    if j ==0:  
#        
#        ws.add_chart(chart,cell(j))
#        #print(cell(j))
#        count = count +18
    
    
    # calcualtes the average of each column
    
    
    for i in range(3,39):
        c2 = LineChart()
        c2.title = switch_demo(i-1)
        c2.style = 12
        c2.y_axis.title = "Settling time (ms)"
        c2.y_axis.crossAx = 500
        c2.x_axis = DateAxis(crossAx=100)
        c2.x_axis.title = "Temp,Voltage"
        data = Reference(ws, min_col=i, min_row=21, max_col=i, max_row=21+len(vdd))
        c2.add_data(data, titles_from_data=True)
        dates = Reference(ws, min_col=1, min_row=22, max_col =2,max_row=22+len(vdd)-1)
        c2.set_categories(dates)
        ws.add_chart(c2,cell(i))
    
    wb.save(file_chart)
    
    








#    chart = ScatterChart()
#    chart.style = 1
#    chart.title = 'Scatter Chart'
#    chart.scatterStyle = 'marker'
#    chart.x_axis.title = 'DUT'
#    chart.y_axis.title = 'Settling times'
#    
#    
#    xvalues = Reference(ws, min_col=1, min_row=2, max_row=13)
#    #for j in range(0,54,18):
#    for i in range(2, 20):
#        values = Reference(ws, min_col=i, min_row=1, max_row=13)
#        series = Series(values, xvalues, title_from_data=True)
#        chart.series.append(series)
#    #    if j ==0:  
#    ws.add_chart(chart, "A25")
#    #    elif j == 18:
#    #        ws.add_chart(chart, "L25")
#    ##    else: 
#    #        ws.add_chart(chart, "Z25")
#    
#    
#    chart1 = ScatterChart()
#    chart1.style = 1
#    chart1.title = 'Scatter Chart'
#    chart1.scatterStyle = 'marker'
#    chart1.x_axis.title = 'DUT'
#    chart1.y_axis.title = 'Settling times'
#    
#    for i in range(20,38):
#        values = Reference(ws, min_col=i, min_row=1, max_row=13)
#        series = Series(values, xvalues, title_from_data=True)
#        chart1.series.append(series)
#    #    if j ==0:  
#    ws.add_chart(chart1, "K25")
#    
#    
#    chart2 = ScatterChart()
#    chart2.style = 1
#    chart2.title = 'Scatter Chart'
#    chart2.scatterStyle = 'marker'
#    chart2.x_axis.title = 'DUT'
#    chart2.y_axis.title = 'Settling times'
#    
#    for i in range(38,56):
#        values = Reference(ws, min_col=i, min_row=1, max_row=13)
#        series = Series(values, xvalues, title_from_data=True)
#        chart2.series.append(series)
#    #    if j ==0:  
#    ws.add_chart(chart2, "V25")

    
    
    

#for i in range(1,13,1):
#    temp_obj=ws.cell(row=i,column=1)