# -*- coding: utf-8 -*-
"""
Created on Fri Jul  6 13:47:50 2018

@author: test
#"""

def switch_demo(argument):
        switcher = {
            1: 'DUT',
            2:"regreadfromsleep",
            3:"regwritefromsleep",
            4: 'regreadfromgyrodisable',
            5:"regwritefromgyrodisable",
            6:"regreadfromacceldisable",
            7: 'regwritefromacceldisable',
            8: 'otpreadfrompowercycle',
            9: 'regreadfrompowercycle',
            10: 'regwritefrompowercycle' ,
        }
        return switcher.get(argument)

def excel_modf(file1,summary_file,UL,LL,vdd):

    from openpyxl import load_workbook
    from openpyxl import Workbook
    import numpy as np
    from statistics import mean
    
    wb = load_workbook(filename=file1, read_only=False)
    ws = wb['Sheet']
    
    data = Workbook()
    datas = data.active
    upperlimit = UL
    lowerlimit = LL
    VDD = len(vdd)
    #define number of DUT = 12 
    
    #filtering out bad data and saving the whole array 
    for i in range(1,ws.max_column+1):
        for j in range(1,ws.max_row+1):
            cell_obj=ws.cell(row=j,column=i)
            if cell_obj.value >= upperlimit or cell_obj.value < lowerlimit:
                datas.cell(row =j , column = i).value = np.nan
            else:
                datas.cell(row =j , column = i).value = cell_obj.value
    
    avg = data.create_sheet('Average')
    avg1 = data.create_sheet('Average1')
    array=[]
    mean1=[]
    
    #for averaging the array and save it 
    #1st for loop isfor DUT
    #2nd for loopis append all the averages of 1DUT 
    
    for z in range(1,datas.max_row+1,1):
        #print("done with one row")
        for i in range(1,13,1):
            #print("done with one one unit")
            array = []
            for j in range(i,datas.max_column+1,12):
                cell_obj=datas.cell(row=z,column=j)
                array.append(cell_obj.value)
            #print(array)
            avg.append(array) # appending 1 DUT array to excel sheet
            mean1.append(np.average(array))# calculate mean including all zeroes. 
    avg1.append(mean1)
    avg2 = data.create_sheet('summary_table')
    
    
    # for inserting the DUT numbers
    # 14 should be replaced by number of DUTs
    #writing DUT num
    for i in range(4,16):
        temp_obj=avg2.cell(row=i,column=1)
        temp_obj.value = i-3
    
    
    
    
    # 9 is the number of conditions 
    #36 = number of conditions (gyro+accel) to lookup from dictionary
    for i in range(0,ws.max_row,9):
        for j in range(2,11):
            temp_obj = avg2.cell(row = 3,column = j+i)
            temp1 = switch_demo(j)
            temp_obj.value=temp1
    obj = avg2.cell(row =3, column =1)
    obj.value = 'DUT'
    
    #for arranging the results in matrix
    k=2
    for i in range(0,avg1.max_column,12):
        for j in range(1,13):
            temp1 = avg1.cell(row=1,column =j+i)
            temp2= avg2.cell(row=j+3,column=k)
            temp2.value = temp1.value
        k=k+1
    
    # for merging the cells for VDD and TEMP
    # total 36 conditons per voltage
#    include when evrything is done
#    for i in range(2,ws.max_column,36):
#        ws.merge_cells('B[i]:[i+37]')
    
    # for averaging the each condition
    mean1 = []
    avg_num = float(1)
    for j in range(2,avg2.max_column+1):
        #print(j)
        for k in range(4,16):
            obj = avg2.cell(row =k, column = j)
            #print(obj.value)
            mean1.append(obj.value)
            print(mean1)
        avg_num = np.nanmean(mean1)
        #print(avg_num)
        #np.nanmean
#        print(avg)
        obj1 = avg2.cell(row = 18, column = j)
        obj1.value =avg_num
        mean1 = []
    
    # 22 and 25 shoud be replaced with variable
    # 18 is equivalent to 18 inabove code snippet 
    # row shodu be set at 18 
    avg_row = 18
    print_row = 22  # this is where the plot table will be printed 
    col = 0
    for j in range(print_row,print_row+VDD,1):
        #for col in range(0,avg2.max_column,36):
            col_1 = 3
            for i in range(2+col,11+col,1):
                obj = avg2.cell(row = avg_row,column = i)
                obj_1 = avg2.cell(row = j, column = col_1)
                #print(obj.value)
                #print(obj_1.value)
                obj_1.value = obj.value
                col_1 = col_1+1
            col = col+9
    
    #for writing the voltages 
    for k in range(0,len(vdd)):
        obj = avg2.cell(row=print_row,column =2 )
        obj.value=vdd[k]
        print_row = print_row+1
    
    #fro writing the each condition in excel sheet for columnn heading
    #for i in range(0,ws.max_row,36):
    
    for j in range(2,11):
        temp_obj = avg2.cell(row = 21,column = j+1)
        temp1 = switch_demo(j)
        temp_obj.value=temp1  
        
        
    data.save(summary_file)
    #data.save("result_1125am.xlsx")
    return 